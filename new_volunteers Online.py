from __future__ import print_function
from apiclient import discovery
from httplib2 import Http
from oauth2client import file, client, tools
from apiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import io
import pyautogui
import subprocess
import webbrowser
import sys
import pyautogui
import time
import smtplib
import openpyxl
from openpyxl import load_workbook

SCOPES = 'https://www.googleapis.com/auth/drive.readonly'
def main():
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('drive', 'v3', http=creds.authorize(Http()))
    file_id = '***'
    request = service.files().export_media(fileId=file_id,
                                                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fh = io.FileIO('volunteerapps.xlsx', 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        print('Volunteer log updated.')
        print('checking for new applicants...')
if __name__ == '__main__':
                    main()
wb = openpyxl.load_workbook('volunteerapps.xlsx')
ws = wb.active
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Spring 2019')
for row in range(2, sheet.max_row +1):
    if sheet['I' + str(row)].value is not None and "DJ" in sheet['I' + str(row)].value and sheet['G' + str(row)].value not in open("PriorDJs.txt").read():
            print('--------------------')
            print(sheet['G' + str(row)].value)
            fromaddr = "[My email]"
            toaddr = sheet['G' + str(row)].value
            msg = MIMEMultipart("alternative")
            msg['From'] = fromaddr
            msg['To'] = toaddr
            msg['Subject'] = "DJ Volunteering with Impact 89FM"
            body = """\
            <html>
                <head></head>
                <body>
                <p>Hello!<br>
                    <br>
                    If you’re receiving this email, you’ve indicated to someone that you’re interested in becoming a DJ at the Impact! Great! Take a moment to skim through the attached DJ handbook and if you’re still interested, fill out <a href="https://goo.gl/forms/ErWwcJYvWUAdQrac2">this form</a> to sign up for training. We’re excited to meet you!<br>
                            <br>
                            <[HTML for email signature]>
                            </body>
            </html>
            """
            msg.attach(MIMEText(body, 'html'))
            filename = "WDBM Airstaff Handbook 2018 .pdf"
            attachment = open("WDBM Airstaff Handbook 2018 .pdf", "rb")
            part = MIMEBase('application', 'octet-stream')
            part.set_payload((attachment).read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
            msg.attach(part)
            server = smtplib.SMTP('smtp-mail.outlook.com', 587)
            server.starttls()
            server.login(fromaddr, "[PASSWORD]")
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()
            open("PriorDJs.txt", "a").write(sheet['G' + str(row)].value)
            open("PriorDJs.txt", "a").close()
            print('--------------------')
            print("New volunteers contacted.")

