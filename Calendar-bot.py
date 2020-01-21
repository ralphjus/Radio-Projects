
from __future__ import print_function
import datetime
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from apiclient import discovery
from httplib2 import Http
from oauth2client import file, client, tools
from apiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
import random
import os
import io
import subprocess
import sys
import pyautogui
import time
import openpyxl
from openpyxl import load_workbook
import re

#Downlaods Calendar as excel
SCOPES = 'https://www.googleapis.com/auth/drive.readonly'
def main():
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('drive', 'v3', http=creds.authorize(Http()))
    file_id = '[ID]'
    request = service.files().export_media(fileId=file_id,
                                              mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fh = io.FileIO('Cal.xlsx', 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
if __name__ == '__main__':
                    main()

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/calendar']

def main():
    """Shows basic usage of the Google Calendar API.
    Prints the start and name of the next 10 events on the user's calendar.
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server()
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('calendar', 'v3', credentials=creds)

    # Call the Calendar API for FIX studio (studio K)
    now = datetime.datetime.utcnow().isoformat() + 'Z' # 'Z' indicates UTC time
    events_result = service.events().list(calendarId='[ID]', timeMin=now,
                                        maxResults=5000, singleEvents=True,
                                        orderBy='startTime').execute()
    events = events_result.get('items', [])

    #Scans FIX scheudle
    #Deletes FIXers no longer on Sheet schedule from Google Calendar
    wb = openpyxl.load_workbook('Cal.xlsx')
    ws = wb.active
    wb.get_sheet_names()
    sheet = wb.get_sheet_by_name('FIX')
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=60, max_col=7):
        for cell in row:
            if cell.value is not None:
                open("FIXCalTxt.txt", "a").write(cell.value)
    for event in events:
        if event['summary'] not in open("FIXCalTxt.txt").read():
            service.events().delete(calendarId='[ID]', eventId= event['id']).execute()
        open("FIXevents.txt", "a").write(event['summary'])
    
    #FIX adds
    def char_range(c1, c2):
        """Generates the characters from `c1` to `c2`, inclusive."""
        for c in range(ord(c1), ord(c2)+1):
            yield chr(c)
    for c in char_range('A', 'G'):
        for row in range(1, 14):
            if sheet[str(c) + str(row*4)].value is not None and sheet[str(c) + str(row*4)].value not in open("FIXevents.txt").read():
                def next_weekday(d, weekday):
                    days_ahead = weekday - d.weekday()
                    if days_ahead <= 0: # Target day already happened this week
                        days_ahead += 7
                    return d + datetime.timedelta(days_ahead)
                d = datetime.datetime.now()
                next_monday = next_weekday(d, (int(ord(c)-65))) # 0 = Monday, 1=Tuesday, 2=Wednesday...
                m = next_monday.strftime("%Y-%m-%d")
                starttime = m + "T" + str(datetime.time(row+7))
                endtime = m + "T" + str(datetime.time(row+8))
                event = {
                'summary': sheet[str(c) + str(row*4)].value,
                'location': '',
                'description': 'Created by Rog Smog.',
                'start': {
                'dateTime': starttime,
                'timeZone': 'America/New_York',
                },
                'end': {
                'dateTime': endtime,
                'timeZone': 'America/New_York',
                },
                'recurrence': [
                'RRULE:FREQ=WEEKLY;COUNT=20'
                ],
            }   
                event = service.events().insert(calendarId='[ID]', body=event).execute()
        #Ends at midnight special case
        if sheet[str(c) + '60'].value is not None and sheet[str(c) + '60'].value not in open("FIXevents.txt").read():
            def next_weekday(d, weekday):
                    days_ahead = weekday - d.weekday()
                    if days_ahead <= 0: # Target day already happened this week
                        days_ahead += 7
                    return d + datetime.timedelta(days_ahead)
            d = datetime.datetime.now()
            current_day = next_weekday(d, (int(ord(c)-65))) # 0 = Monday, 1=Tuesday, 2=Wednesday...
            next_day = next_weekday(d, (int(ord(c)-64)))
            m = current_day.strftime("%Y-%m-%d")
            t = next_day.strftime("%Y-%m-%d")
            starttime = m + "T" + "22:00:00-04:00"
            endtime = t + "T" + "00:00:00-04:00"
            event = {
            'summary': sheet[str(c) + '60'].value,
            'location': '',
            'description': 'Created by Rog Smog.',
            'start': {
            'dateTime': starttime,
            'timeZone': 'America/New_York',
            },
            'end': {
            'dateTime': endtime,
            'timeZone': 'America/New_York',
            },
            'recurrence': [
            'RRULE:FREQ=WEEKLY;COUNT=20'
            ],
        }
            event = service.events().insert(calendarId='[ID]', body=event).execute()
    open("FIXCalTxt.txt", "w+").close()
    open("FIXevents.txt", "w+").close()

    # Call the Calendar API for WDBM studio (Studio I)
    now = datetime.datetime.utcnow().isoformat() + 'Z' # 'Z' indicates UTC time
    events_result = service.events().list(calendarId='[ID]', timeMin=now,
                                        maxResults=1, singleEvents=True,
                                        orderBy='startTime').execute()
    events = events_result.get('items', [])

    #Scans WDBM scheudle
    #Deletes DJs no longer on Sheet schedule from Google Calendar
    wb = openpyxl.load_workbook('Cal.xlsx')
    ws = wb.active
    wb.get_sheet_names()
    sheet = wb.get_sheet_by_name('New WDBM')
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=32, max_col=7):
        for cell in row:
            if cell.value is not None:
                open("WDBMCalTxt.txt", "a").write(cell.value)
    for event in events:
        if event['summary'] not in open("WDBMCalTxt.txt").read():
            service.events().delete(calendarId='[ID]', eventId= event['id']).execute()
        open("WDBMevents.txt", "a").write(event['summary'])
    
    #WDBM adds
    def char_range(c1, c2):
        """Generates the characters from `c1` to `c2`, inclusive."""
        for c in range(ord(c1), ord(c2)+1):
            yield chr(c)
    for c in char_range('A', 'G'):
        for row in range(1, 7):
            if sheet[str(c) + str(row*4)].value is not None and sheet[str(c) + str(row*4)].value not in open("WDBMevents.txt").read():
                def next_weekday(d, weekday):
                    days_ahead = weekday - d.weekday()
                    if days_ahead <= 0: # Target day already happened this week
                        days_ahead += 7
                    return d + datetime.timedelta(days_ahead)
                d = datetime.datetime.now()
                current_day = next_weekday(d, (int(ord(c)-65))) # 0 = Monday, 1=Tuesday, 2=Wednesday...
                next_day = next_weekday(d, (int(ord(c)-64)))
                m = current_day.strftime("%Y-%m-%d")
                t = next_day.strftime("%Y-%m-%d")
                starttime = m + "T" + str(datetime.time((2*row)+6))
                endtime = m + "T" + str(datetime.time((2*row)+8))
                midnight = t + "T" + "00:00"
                event = {
                'summary': sheet[str(c) + str(row*4)].value,
                'location': '',
                'description': 'Created by Rog Smog.',
                'start': {
                'dateTime': starttime,
                'timeZone': 'America/New_York',
                },
                'end': {
                'dateTime': endtime,
                'timeZone': 'America/New_York',
                },
                'recurrence': [
                'RRULE:FREQ=WEEKLY;COUNT=20'
                ],
            }
                event = service.events().insert(calendarId='[ID]', body=event).execute()
        #Ends at midnight special case
        if sheet[str(c) + '32'].value is not None and sheet[str(c) + '32'].value not in open("WDBMevents.txt").read():
            def next_weekday(d, weekday):
                    days_ahead = weekday - d.weekday()
                    if days_ahead <= 0: # Target day already happened this week
                        days_ahead += 7
                    return d + datetime.timedelta(days_ahead)
            d = datetime.datetime.now()
            current_day = next_weekday(d, (int(ord(c)-65))) # 0 = Monday, 1=Tuesday, 2=Wednesday...
            next_day = next_weekday(d, (int(ord(c)-64)))
            m = current_day.strftime("%Y-%m-%d")
            t = next_day.strftime("%Y-%m-%d")
            starttime = m + "T" + "22:00:00-04:00"
            endtime = t + "T" + "00:00:00-04:00"
            event = {
            'summary': sheet[str(c) + '32'].value,
            'location': '',
            'description': 'Created by Rog Smog.',
            'start': {
            'dateTime': starttime,
            'timeZone': 'America/New_York',
            },
            'end': {
            'dateTime': endtime,
            'timeZone': 'America/New_York',
            },
            'recurrence': [
            'RRULE:FREQ=WEEKLY;COUNT=20'
            ],
        }
            event = service.events().insert(calendarId='[ID]', body=event).execute()
    open("WDBMCalTxt.txt", "w+").close()
    open("WDBMevents.txt", "w+").close()
    

if __name__ == '__main__':
    main()
