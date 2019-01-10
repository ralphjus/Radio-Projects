from __future__ import print_function
from apiclient import discovery
from httplib2 import Http
from oauth2client import file, client, tools
from apiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import random
import os
import io
import pyautogui
import subprocess
import webbrowser
import sys
import pyautogui
import time
import smtplib
import openpyxl
from openpyxl import load_workbook
import re
from slackclient import SlackClient

#Downlaods training sign up
SCOPES = 'https://www.googleapis.com/auth/drive.readonly'
def main():
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('drive', 'v3', http=creds.authorize(Http()))
    file_id = '***'
    request = service.files().export_media(fileId=file_id,
                                              mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fh = io.FileIO('FIXsignups.xlsx', 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        print("FIXsignups.xlsx updated.")
if __name__ == '__main__':
                    main()
#Downloads Personnel database
def main():
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('drive', 'v3', http=creds.authorize(Http()))
    file_id = '***'
    request = service.files().export_media(fileId=file_id,
                                              mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fh = io.FileIO('Database.xlsx', 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        print('Database.xlsx updated.')
if __name__ == '__main__':
                    main()
#Downloads Feedback Form
def main():
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', SCOPES)
        creds = tools.run_flow(flow, store)
    service = build('drive', 'v3', http=creds.authorize(Http()))
    file_id = '***'
    request = service.files().export_media(fileId=file_id,
                                              mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fh = io.FileIO('Feedback_Form.xlsx', 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        print('Feedback_Form.xlsx updated.')
if __name__ == '__main__':
                    main()
                    
# instantiate Slack client
slack_client = SlackClient(os.environ.get('SLACK_BOT_TOKEN'))
# starterbot's user ID in Slack: value is assigned after the bot starts up
starterbot_id = None
# constants
RTM_READ_DELAY = 1 # 1 second delay between reading from RTM
MENTION_REGEX = "^<@(|[WU].+?)>(.*)"

#Notifies JUSTIN iff FIX trainees haven't been contacted
wb = openpyxl.load_workbook('FIXsignups.xlsx')
ws = wb.active
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Form Responses 1')
print("scanning for needy FIX trainees...")
for row in range(2, sheet.max_row +1):
      if sheet['I' + str(row)].value is None:
          slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['This trainee needs to be contacted!', sheet['C' + str(row)].value, sheet['B' + str(row)].value]
        )

#Sends DM to JUSTIN iff volunteer available
wb = openpyxl.load_workbook('FIXsignups.xlsx')
ws = wb.active
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Form Responses 1')
print("checking for new FIX applicants...")
for row in range(2, sheet.max_row +1):
    if sheet['D' + str(row)].value is not None and "Justin" in sheet['D' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainJUSM.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "MONDAYS", sheet['D' + str(row)].value]
        )
        open("PriorTrainJUSM.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainJUSM.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Justin.')
    if sheet['E' + str(row)].value is not None and "Justin" in sheet['E' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainJUST.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "TUESDAYS", sheet['E' + str(row)].value]
        )
        open("PriorTrainJUST.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainJUST.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Justin')
    if sheet['F' + str(row)].value is not None and "Justin" in sheet['F' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainJUSW.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "WEDNESDAYS", sheet['F' + str(row)].value]
        )
        open("PriorTrainJUSW.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainJUSW.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Justin.')
    if sheet['G' + str(row)].value is not None and "Justin" in sheet['G' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainJUSTH.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "THURSDAYS", sheet['G' + str(row)].value]
        )
        open("PriorTrainJUSTH.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainJUSTH.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Justin.')
    if sheet['H' + str(row)].value is not None and "Justin" in sheet['H' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainJUSF.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "FRIDAYS", sheet['G' + str(row)].value]
        )
        open("PriorTrainJUSF.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainJUSF.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Justin.')
        
#Sends DM to ZOE iff volunteer available
for row in range(2, sheet.max_row +1):
    if sheet['D' + str(row)].value is not None and "Zoe" in sheet['D' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainZOEM.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "MONDAYS", sheet['D' + str(row)].value]
        )
        open("PriorTrainZOEM.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainZOEM.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Zoe.')
    if sheet['E' + str(row)].value is not None and "Zoe" in sheet['E' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainZOET.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "TUESDAYS", sheet['E' + str(row)].value]
        )
        open("PriorTrainZOET.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainZOET.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Zoe.')
    if sheet['F' + str(row)].value is not None and "Zoe" in sheet['F' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainZOEW.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "WEDNESDAYS", sheet['F' + str(row)].value]
        )
        open("PriorTrainZOEW.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainZOEW.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Zoe.')
    if sheet['G' + str(row)].value is not None and "Zoe" in sheet['G' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainZOETH.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "THURSDAYS", sheet['G' + str(row)].value]
        )
        open("PriorTrainZOETH.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainZOETH.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Zoe.')
    if sheet['H' + str(row)].value is not None and "Zoe" in sheet['H' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainZOEF.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "FRIDAYS", sheet['G' + str(row)].value]
        )
        open("PriorTrainZOEF.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainZOEF.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Zoe.')
        
#Sends DM to TYLER iff volunteer available
for row in range(2, sheet.max_row +1):
    if sheet['D' + str(row)].value is not None and "Tyler" in sheet['D' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainTYLM.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "MONDAYS", sheet['D' + str(row)].value]
        )
        open("PriorTrainTYLM.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainTYLM.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Tyler.')
    if sheet['E' + str(row)].value is not None and "Tyler" in sheet['E' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainTYLT.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "TUESDAYS", sheet['E' + str(row)].value]
        )
        open("PriorTrainTLYT.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainTYLT.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Tyler.')
    if sheet['F' + str(row)].value is not None and "Tyler" in sheet['F' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainTYLW.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "WEDNESDAYS", sheet['F' + str(row)].value]
        )
        open("PriorTrainTYLW.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainTYLW.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Tyler.')
    if sheet['G' + str(row)].value is not None and "Tyler" in sheet['G' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainTYLTH.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "THURSDAYS", sheet['G' + str(row)].value]
        )
        open("PriorTrainTYLTH.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainTYLTH.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Tyler.')
    if sheet['H' + str(row)].value is not None and "Tyler" in sheet['H' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainTYLF.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "FRIDAYS", sheet['H' + str(row)].value]
        )
        open("PriorTrainTYLF.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainTYLF.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Tyler.')

#Sends DM to AMANDA iff volunteer available
for row in range(2, sheet.max_row +1):
    if sheet['D' + str(row)].value is not None and "Amanda" in sheet['D' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainAMAM.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "MONDAYS", sheet['D' + str(row)].value]
        )
        open("PriorTrainAMAM.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainAMAM.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Amanda.')
    if sheet['E' + str(row)].value is not None and "Amanda" in sheet['E' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainAMAT.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "TUESDAYS", sheet['E' + str(row)].value]
        )
        open("PriorTrainAMAT.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainAMAT.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Amanda.')
    if sheet['F' + str(row)].value is not None and "Amanda" in sheet['F' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainAMAW.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "WEDNESDAYS", sheet['F' + str(row)].value]
        )
        open("PriorTrainAMAW.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainAMAW.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Amanda.')
    if sheet['G' + str(row)].value is not None and "Amanda" in sheet['G' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainAMATH.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "THURSDAYS", sheet['G' + str(row)].value]
        )
        open("PriorTrainAMATH.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainAMATH.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Amanda.')
    if sheet['H' + str(row)].value is not None and "Amanda" in sheet['H' + str(row)].value and sheet['B' + str(row)].value not in open("PriorTrainAMAF.txt").read():
        slack_client.api_call(
            "chat.postMessage",
            channel="***",
            text= ['A new volunteer wants to train with you!', sheet['C' + str(row)].value, sheet['B' + str(row)].value, "FRIDAYS", sheet['G' + str(row)].value]
        )
        open("PriorTrainAMAF.txt", "a").write(sheet['B' + str(row)].value)
        open("PriorTrainAMAF.txt", "a").close()
        print(sheet['B' + str(row)].value, 'sent to Amanda.')


def parse_bot_commands(slack_events):
    """
        Parses a list of events coming from the Slack RTM API to find bot commands.
        If a bot command is found, this function returns a tuple of command and channel.
        If its not found, then this function returns None, None.
    """
    for event in slack_events:
        if event["type"] == "message" and not "subtype" in event:
            user_id, message = parse_direct_mention(event["text"])
            if user_id == starterbot_id:
                return message, event["channel"]
    return None, None

def parse_direct_mention(message_text):
    """
        Finds a direct mention (a mention that is at the beginning) in message text
        and returns the user ID which was mentioned. If there is no direct mention, returns None
    """
    matches = re.search(MENTION_REGEX, message_text)
    # the first group contains the username, the second group contains the remaining message
    return (matches.group(1), matches.group(2).strip()) if matches else (None, None)

def handle_command(command, channel):
    """
        Executes bot command if the command is known
    """
    # Default response is help text for the user
    default_response = "Person not found. Last name may need to be capitalized. Try *{}*.".format("@Rog Smog find [Last name]")

    # Finds and executes the given command, filling in response
    response = None

    #Score lookup
    if command.startswith("score"):
        print("Score request recieved. Processing...")
        wb = openpyxl.load_workbook('Feedback_Form.xlsx')
        sheet = wb.active
        for row in range(2, sheet.max_row +1):
            if sheet['B' + str(row)].value is not None:
                open("FeedbackEmails.txt", "a").write(sheet['B' + str(row)].value)
                open("PriorTrainJUST.txt", "a").close()
                if sheet['B' + str(row)].value is not None and eval('command').replace('score ','') in sheet['B' + str(row)].value:
                    slack_client.api_call(
                        "chat.postMessage",
                        channel=channel,
                        text= ["FIX", sheet['BX' + str(row)].value, "VT", sheet['AZ' + str(row)].value, "Live", sheet['AB' + str(row)].value]
                    )
                    print("Score request processed.")
        if eval('command').replace('score ','') not in open("FeedbackEmails.txt").read():
            slack_client.api_call(
                "chat.postMessage",
                channel=channel,
                text= "Email not found."
            )
        open('FeedbackEmails.txt', 'w').close()
                   
    #Person Lookup
    if command.startswith("find"):
        print("Info request recieved. Processing...")
        wb = openpyxl.load_workbook('Database.xlsx')
        sheet = wb.active
        for row in range(2, sheet.max_row +1):
            if sheet['D' + str(row)].value is not None and eval('command').replace('find ','') in sheet['D' + str(row)].value:
                response = [sheet['B' + str(row)].value, sheet['D' + str(row)].value,  sheet['G' + str(row)].value,   sheet['L' + str(row)].value,  sheet['H' + str(row)].value,  sheet['R' + str(row)].value]
        # Sends the response back to the channel
        slack_client.api_call(
            "chat.postMessage",
            channel=channel,
            text=response or default_response
        )
        print("request processed.")
        print(eval('command').replace('find ',''))

if __name__ == "__main__":
    if slack_client.rtm_connect(with_team_state=False):
        print("Slack Bot is online.")
        # Read bot's user ID by calling Web API method `auth.test`
        starterbot_id = slack_client.api_call("auth.test")["user_id"]
        while True:
            command, channel = parse_bot_commands(slack_client.rtm_read())
            if command:
                handle_command(command, channel)
            time.sleep(RTM_READ_DELAY)
    else:
        print("Connection failed. Exception traceback printed above.")
